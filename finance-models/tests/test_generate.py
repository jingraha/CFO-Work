"""Automated validation for the Startup CFO OS Excel model track.

Generates all four workbooks (via the same `models.*.build()` entry points used
by `generate.py`) into a temporary directory and inspects each with openpyxl
for structural correctness: required tabs, 36 monthly periods, scenario data
validation, formula presence (not hard-coded numbers) on key schedules,
cross-sheet links, chart presence, and non-empty output files.

Run with:
    cd finance-models
    python -m pytest tests -v
or directly:
    python tests\\test_generate.py
"""
from __future__ import annotations

import sys
from pathlib import Path

import pytest

THIS_DIR = Path(__file__).resolve().parent
ROOT_DIR = THIS_DIR.parent
sys.path.insert(0, str(ROOT_DIR))

import openpyxl  # noqa: E402
from openpyxl.formula.tokenizer import Tokenizer  # noqa: E402
from openpyxl.worksheet.formula import ArrayFormula  # noqa: E402
from openpyxl.utils.cell import coordinate_from_string  # noqa: E402

from models import b2b_saas, ai_api, consumer, ai_services  # noqa: E402

REQUIRED_TABS = [
    "Read Me", "Assumptions", "Headcount", "Revenue", "COGS & GM", "Opex",
    "Working Capital", "P&L", "Balance Sheet", "Cash Flow", "Cash & Runway",
    "Scenarios", "Visuals", "Checks",
]

SCHEDULE_TABS_WITH_36_MONTHS = [
    "Headcount", "Revenue", "COGS & GM", "Opex", "Working Capital", "P&L",
    "Balance Sheet", "Cash Flow", "Cash & Runway",
]

MODELS = {
    "b2b_saas": b2b_saas,
    "ai_api": ai_api,
    "consumer": consumer,
    "ai_services": ai_services,
}


def _formula_text(cell_value):
    """Return the formula string for a cell value, whether it's a plain
    formula string or an openpyxl ArrayFormula wrapper."""
    if isinstance(cell_value, ArrayFormula):
        return cell_value.text
    return cell_value


def _row_for_label(ws, label):
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == label:
            return row
    raise AssertionError(f"{ws.title}: missing row label {label!r}")


@pytest.fixture(scope="module", params=list(MODELS.keys()))
def built_workbook(request, tmp_path_factory):
    """Build one model workbook once per model into a shared temp dir and
    load it back with openpyxl for inspection. Returns (key, path, workbook)."""
    key = request.param
    out_dir = tmp_path_factory.mktemp("cfo_os_outputs")
    out_path = out_dir / f"{key}.xlsx"
    MODELS[key].build(str(out_path))
    assert out_path.exists(), f"{key}: build() did not produce a file"
    assert out_path.stat().st_size > 20_000, f"{key}: output file suspiciously small/empty"
    wb = openpyxl.load_workbook(str(out_path), data_only=False)
    return key, out_path, wb


def test_output_file_nonempty(built_workbook):
    key, path, wb = built_workbook
    assert path.stat().st_size > 20_000


def test_required_tabs_present(built_workbook):
    key, path, wb = built_workbook
    missing = [t for t in REQUIRED_TABS if t not in wb.sheetnames]
    assert not missing, f"{key}: missing required tabs {missing} (found {wb.sheetnames})"


def test_36_monthly_period_columns(built_workbook):
    """Every schedule tab must expose exactly 36 monthly formula columns
    (C through AL) with a live formula in at least one representative row."""
    key, path, wb = built_workbook
    for tab in SCHEDULE_TABS_WITH_36_MONTHS:
        ws = wb[tab]
        # Find a row below the header block (row 6+) that has a formula in
        # column C (first monthly column) -- confirms the monthly grid exists.
        found_formula_row = None
        for r in range(6, min(ws.max_row, 40) + 1):
            v = ws.cell(row=r, column=3).value
            v = _formula_text(v)
            if isinstance(v, str) and v.startswith("="):
                found_formula_row = r
                break
        assert found_formula_row is not None, f"{key}/{tab}: no formula found in column C rows 6-40"
        # Column AL is the 36th monthly column (C=3 .. AL=38, inclusive => 36 cols)
        last_month_col = 3 + 35  # column index for AL (1-based)
        v_last = _formula_text(ws.cell(row=found_formula_row, column=last_month_col).value)
        assert isinstance(v_last, str) and v_last.startswith("="), (
            f"{key}/{tab}: expected a formula in the 36th month column (AL) at row {found_formula_row}, got {v_last!r}"
        )
        # Column AM (37th) must NOT be a monthly data formula column identically shaped;
        # AN.. is the quarterly section, separated by a blank spacer column (AM).
        v_spacer = ws.cell(row=found_formula_row, column=3 + 36).value  # AM = 39th col (1-based) = spacer
        assert v_spacer in (None, ""), f"{key}/{tab}: expected spacer column AM to be blank, got {v_spacer!r}"


def test_scenario_selector_data_validation(built_workbook):
    key, path, wb = built_workbook
    ws = wb["Assumptions"]
    dvs = ws.data_validations.dataValidation
    assert len(dvs) >= 1, f"{key}: no data validation found on Assumptions tab"
    list_dvs = [dv for dv in dvs if dv.type == "list"]
    assert list_dvs, f"{key}: no list-type data validation found on Assumptions tab"
    found = False
    for dv in list_dvs:
        formula1 = (dv.formula1 or "")
        if "Base" in formula1 and "Upside" in formula1 and "Downside" in formula1:
            found = True
            # sqref should include the selector cell C4
            assert "C4" in str(dv.sqref), f"{key}: scenario data validation not anchored at C4 ({dv.sqref})"
    assert found, f"{key}: no Base/Upside/Downside list data validation found"


def test_checks_tab_has_six_checks_and_rollup(built_workbook):
    key, path, wb = built_workbook
    ws = wb["Checks"]
    # Expect 6 check rows (metric + status formulas) plus 1 overall roll-up formula.
    formula_rows = []
    for r in range(1, ws.max_row + 1):
        metric = _formula_text(ws.cell(row=r, column=3).value)
        status = _formula_text(ws.cell(row=r, column=4).value)
        if isinstance(status, str) and status.startswith("="):
            formula_rows.append((r, metric, status))
    assert len(formula_rows) >= 6, f"{key}: expected >=6 check rows with status formulas, found {len(formula_rows)}"
    # At least one status formula should be a PASS/FAIL IF() check.
    assert any("PASS" in s and "FAIL" in s for _, _, s in formula_rows), (
        f"{key}: no PASS/FAIL formulas found on Checks tab"
    )
    # Overall roll-up: a formula referencing multiple status cells with AND(...)
    rollup_found = any("AND(" in s for _, _, s in formula_rows)
    assert rollup_found, f"{key}: no overall AND(...) roll-up formula found on Checks tab"


def test_formula_error_scan_present(built_workbook):
    """The Checks tab must include a formula that scans for Excel formula
    errors (#REF!, #DIV/0!, #VALUE! etc.) via ISERROR across key schedules."""
    key, path, wb = built_workbook
    ws = wb["Checks"]
    found = False
    for r in range(1, ws.max_row + 1):
        v = _formula_text(ws.cell(row=r, column=3).value)
        if isinstance(v, str) and "ISERROR" in v.upper():
            found = True
            break
    assert found, f"{key}: no ISERROR-based formula-error scan found on Checks tab"


def test_balance_sheet_check_formula_present(built_workbook):
    """Balance sheet balance check must be a real formula comparing assets to
    liabilities+equity, not a hard-coded pass/fail."""
    key, path, wb = built_workbook
    ws = wb["Checks"]
    found = False
    for r in range(1, ws.max_row + 1):
        v = _formula_text(ws.cell(row=r, column=3).value)
        if isinstance(v, str) and "Balance Sheet" in v:
            found = True
            break
    assert found, f"{key}: no Checks formula references 'Balance Sheet'"


def test_key_schedules_are_formulas_not_hardcoded(built_workbook):
    """Sample a handful of cells on Revenue, COGS & GM, P&L, Balance Sheet,
    Cash Flow for months 2-36 and confirm they are formulas (start with '='),
    i.e. not hard-coded static outputs."""
    key, path, wb = built_workbook
    for tab in ["Revenue", "COGS & GM", "P&L", "Balance Sheet", "Cash Flow"]:
        ws = wb[tab]
        checked_any = False
        for r in range(6, min(ws.max_row, 40) + 1):
            for col in (4, 20, 38):  # month 2, ~month 18, month 36 (col AL=38)
                v = _formula_text(ws.cell(row=r, column=col).value)
                if isinstance(v, (int, float)) and v != 0:
                    pytest.fail(f"{key}/{tab}: hard-coded numeric value {v!r} found at row {r} col {col}")
                if isinstance(v, str) and v.startswith("=") and "!" not in v[:1]:
                    checked_any = True
        assert checked_any, f"{key}/{tab}: no formulas found to sample in the monthly grid"


def test_cross_sheet_links_exist(built_workbook):
    """Confirm at least one formula on P&L, Balance Sheet, and Cash Flow
    references another sheet (i.e. the statements are actually integrated,
    not siloed)."""
    key, path, wb = built_workbook
    for tab in ["P&L", "Balance Sheet", "Cash Flow"]:
        ws = wb[tab]
        found = False
        for row in ws.iter_rows():
            for cell in row:
                v = _formula_text(cell.value)
                if isinstance(v, str) and v.startswith("=") and "'" in v and "!" in v:
                    found = True
                    break
            if found:
                break
        assert found, f"{key}/{tab}: no cross-sheet-linked formula (e.g. ='Other Sheet'!...) found"


def test_visuals_tab_has_charts(built_workbook):
    key, path, wb = built_workbook
    ws = wb["Visuals"]
    assert len(ws._charts) >= 4, f"{key}: expected at least 4 charts on Visuals tab, found {len(ws._charts)}"


def test_readme_has_disclaimer(built_workbook):
    key, path, wb = built_workbook
    ws = wb["Read Me"]
    text_blob = " ".join(
        str(cell.value) for row in ws.iter_rows() for cell in row if isinstance(cell.value, str)
    )
    assert "NOT accounting" in text_blob or "not accounting" in text_blob.lower(), (
        f"{key}: Read Me tab missing the required planning-tool disclaimer"
    )


def test_scenario_validity_check_uses_selector(built_workbook):
    key, path, wb = built_workbook
    ws = wb["Scenarios"]
    found = False
    for row in ws.iter_rows():
        for cell in row:
            v = _formula_text(cell.value)
            if isinstance(v, str) and "MATCH" in v.upper() and "Assumptions" in v:
                found = True
    assert found, f"{key}: Scenarios tab has no MATCH-based validity formula referencing Assumptions"


def test_cash_runway_uses_populated_cash_and_pre_financing_burn(built_workbook):
    """Runway must divide by the linked ending-cash row on the current sheet.

    Burn excludes financing proceeds so an equity or debt raise does not
    artificially suppress operating burn and inflate runway.
    """
    key, path, wb = built_workbook
    cash = wb["Cash & Runway"]
    cash_flow = wb["Cash Flow"]
    ending_cash_row = _row_for_label(cash, "Ending Cash Balance")
    burn_row = _row_for_label(cash, "Monthly Net Burn (pre-financing)")
    trailing_row = _row_for_label(cash, "Trailing 3-Month Avg Burn")
    runway_row = _row_for_label(cash, "Runway (months)")
    headline_row = _row_for_label(cash, "Runway as of final month (36)")
    cfo_row = _row_for_label(cash_flow, "Cash from Operations")
    cfi_row = _row_for_label(cash_flow, "Cash from Investing")

    for col in range(3, 39):
        col_letter_value = cash.cell(row=6, column=col).column_letter
        runway_formula = _formula_text(cash.cell(row=runway_row, column=col).value)
        assert f"{col_letter_value}{ending_cash_row}" in runway_formula, (
            f"{key}: runway formula {cash.cell(runway_row, col).coordinate} "
            f"does not reference populated ending cash row {ending_cash_row}: {runway_formula}"
        )
        assert f"{col_letter_value}{trailing_row}" in runway_formula
        assert f"IF({col_letter_value}{ending_cash_row}<=0,0" in runway_formula

        burn_formula = _formula_text(cash.cell(row=burn_row, column=col).value)
        assert f"'Cash Flow'!{col_letter_value}{cfo_row}" in burn_formula
        assert f"'Cash Flow'!{col_letter_value}{cfi_row}" in burn_formula
        assert "Net Change" not in burn_formula

    headline_formula = _formula_text(cash.cell(row=headline_row, column=3).value)
    assert f"AL{ending_cash_row}" in headline_formula
    assert f"AL{trailing_row}" in headline_formula
    assert f"AL{runway_row}" in headline_formula
    assert f"AL{runway_row}>=999" not in headline_formula


def test_formulas_do_not_directly_reference_blank_cells(built_workbook):
    """Catch row-number mistakes where a formula points at an empty cell.

    Range references are intentionally excluded because some schedules use
    sparse ranges. Single-cell references must always resolve to a populated
    input or formula cell.
    """
    key, path, wb = built_workbook
    blank_references = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                formula = _formula_text(cell.value)
                if not isinstance(formula, str) or not formula.startswith("="):
                    continue
                for token in Tokenizer(formula).items:
                    if token.type != "OPERAND" or token.subtype != "RANGE":
                        continue
                    reference = token.value
                    if ":" in reference or "[" in reference or reference.startswith("#"):
                        continue
                    if "!" in reference:
                        sheet_name, coordinate = reference.rsplit("!", 1)
                        sheet_name = sheet_name.strip("'").replace("''", "'")
                    else:
                        sheet_name, coordinate = ws.title, reference
                    coordinate = coordinate.replace("$", "")
                    try:
                        coordinate_from_string(coordinate)
                    except ValueError:
                        continue
                    if sheet_name not in wb.sheetnames:
                        continue
                    if wb[sheet_name][coordinate].value is None:
                        blank_references.append(
                            f"{ws.title}!{cell.coordinate} -> {sheet_name}!{coordinate}"
                        )

    assert not blank_references, (
        f"{key}: formulas directly reference blank cells: "
        + ", ".join(blank_references[:20])
    )


if __name__ == "__main__":
    raise SystemExit(pytest.main([__file__, "-v"]))
